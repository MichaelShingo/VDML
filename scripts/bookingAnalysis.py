import openpyxl
import csv
import os
import pathlib
from datetime import datetime

DATEREGEX = '%Y-%m-%d %H:%M'
DAY = 86400
HOUR = 3600
MINUTE = 60

def sortPair(list1, list2):
    zippedPairs = zip(list2, list1)
    z = [x for _, x in sorted(zippedPairs, reverse=True)]
    return z

def dictToSortedList(dictionary):
    list2 = list(dictionary.keys()) #time difference
    list1 = list(dictionary.values()) #number of reservations
    list2 = sortPair(list2, list1)
    list1.sort(reverse=True)
    return (list2, list1)

def listsToExcel(wsName, list1, list2, column1='Untitled', column2='Untitled', dim1=20, dim2=20):
    for i in range(len(list1)):
        wsName['A' + str(i + 2)] = list1[i]
        wsName['B' + str(i + 2)] = list2[i]
    wsName['A1'] = column1
    wsName['B1'] = column2
    wsName.column_dimensions['A'].width = dim1
    wsName.column_dimensions['B'].width = dim2

def setTotal(wsName, heading, bEndNumber, columnWidth=22):
    wsName['D1'] = heading
    wsName['D2'] = f'=SUM(B2:B{bEndNumber})' 
    wsName.column_dimensions['D'].width = columnWidth
    wsName.column_dimensions['B'].width = columnWidth

#Data retrieval and calculations

def analyzeCSV(filename):

    equipmentToCount = {}
    equipmentBarcodeCount = []
    dayToCount = {i:0 for i in range(7)}
    bookingCount = 0

    hourList = [hour for hour in range(24)]
    hourToCount = {i: 0 for i in hourList}

    dayList = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    dayListMon = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    dayHourList = []
    for day in range(7):
        for hour in hourList:
            dayHourList.append((day, hour))

    dayHourToCount = {x: 0 for x in dayHourList}
    categoryToCount = {}
    noShowToCount = {}
    resToOutTime = {-2:0, -1:0, 0:0} #-2 = immediate, -1 = less than 1 hour, 0 = > 1 hour and < 24 hours
    uniqueCons = set()

    nameToBookings = {} #maps username to #checkouts, #equipment

    nameToLateReturn = [] #maps username to how late equipment was returned, per booking 

    currentDir = pathlib.Path(__file__).parent.resolve()
    os.chdir(os.path.join(currentDir, './../static/files/'))
    rowCount = 0


    with open(filename, 'r', encoding='utf-16') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter='\t')
        next(csv_reader)
        line_count = 0
        for row in csv_reader:
            if row[10] != '':
                currentEquipment = row[10]
            else:
                currentEquipment = row[9]
            if currentEquipment in equipmentToCount:
                equipmentToCount[currentEquipment] += 1
            else:
                equipmentToCount[currentEquipment] = 1
            if not row[7] in categoryToCount:
                categoryToCount[row[7]] = 1
            else:
                categoryToCount[row[7]] += 1
            if row[16] == 'Cancelled: Late checkout policy': #need to look at unique bookings here...not equipment
                lateUser = row[1]
                if not lateUser in noShowToCount:
                    noShowToCount[lateUser] = 1
                else:
                    if not row[0] in uniqueCons:
                        noShowToCount[lateUser] += 1

            if not row[0] in uniqueCons:
                #calculates popular times of the day
                currentDate = datetime.strptime(row[4], DATEREGEX)
                currentDay = currentDate.weekday()
                currentHour = currentDate.hour
                dayTimeTuple = (currentDay, currentHour)
                if not dayTimeTuple in dayHourToCount:
                    dayHourToCount[dayTimeTuple] = 0
                else:
                    dayHourToCount[dayTimeTuple] += 1
                dayToCount[currentDay] += 1
                hourToCount[currentHour] += 1
            bookingCount += 1

            #Time Difference
            if not row[13] == '':
                timeCreated = datetime.strptime(row[3], DATEREGEX)
                timeCheckedOut = datetime.strptime(row[13], DATEREGEX)
                differenceSeconds = (timeCheckedOut - timeCreated).total_seconds() #this works
                differenceDays = int(differenceSeconds // DAY) #this works
            if not row[0] in uniqueCons:
                if differenceDays == 0: #if diff is less than 1 day
                    if differenceSeconds == 0: #immediate checkout (-1)
                        resToOutTime[-2] += 1
                    elif differenceSeconds < HOUR: #greater than 0, less than 1 hour (0)
                        resToOutTime[-1] += 1
                    elif differenceSeconds > HOUR and differenceSeconds < DAY:
                        resToOutTime[0] += 1
                else: #when time difference >= 1
                    if not differenceDays in resToOutTime:
                        resToOutTime[differenceDays] = 1
                    else:
                        resToOutTime[differenceDays] += 1

            #Popular Users - number of checkouts and number of equipment booked
            if not row[1] in nameToBookings:
                nameToBookings[row[1]] = [1, 0] #first is numEquipment, second is numCheckouts
            else:
                nameToBookings[row[1]][0] += 1
            if not row[0] in uniqueCons:
                nameToBookings[row[1]][1] += 1

            #Late Returns

            if not row[15] == '' and not row[0] in uniqueCons:
                checkInTime = datetime.strptime(row[15], DATEREGEX)
                dueTime = datetime.strptime(row[5], DATEREGEX)
                timeDelta = checkInTime - dueTime
                lateString = str(timeDelta.days) + ' Days ' + str(timeDelta.seconds // 3600) + ' Hours ' + str((timeDelta.seconds // 60) % 60) + ' Minutes '
                amountLate = timeDelta.total_seconds()
                if amountLate > 0:
                    nameToLateReturn.append([amountLate, row[1], lateString])
    

            uniqueCons.add(row[0]) #this must be at the end of the loop


    print(len(uniqueCons)) #1985
    print('booking count = ' + str(bookingCount)) #3454

    #Convert dictionaries to lists and sort
    equipmentList = list(equipmentToCount.keys())
    countList = list(equipmentToCount.values())

    hourList = list(hourToCount.keys())
    hourCountList = list(hourToCount.values())

    popularDayCount = list(dayToCount.values())
    popularDayCount.insert(0, popularDayCount[6])
    popularDayCount.pop(7)

    equipmentList = sortPair(equipmentList, countList)
    countList.sort(reverse=True)

    noShowUsername = list(noShowToCount.keys())
    noShowCount = list(noShowToCount.values())
    noShowUsername = sortPair(noShowUsername, noShowCount)
    noShowCount.sort(reverse=True)

    categoryList = list(categoryToCount.keys())
    categoryCount = list(categoryToCount.values())
    categoryList = sortPair(categoryList, categoryCount)
    categoryCount.sort(reverse=True)

    timeDifferenceList, timeDifferenceCount = dictToSortedList(resToOutTime)
    timeDiffToCount = []
    for i in range(len(timeDifferenceCount)):
        timeDiffToCount.append([timeDifferenceList[i], timeDifferenceCount[i]])
    timeDiffToCount.sort()

    #creating and formatting the Excel document
    wb = openpyxl.Workbook()
    wb.active.title = "Number of Bookings by Equipment"
    wsBookingToEquipment = wb.active
    wb.create_sheet('Most Popular Days')
    wb.create_sheet('Most Popular Hours')
    wb.create_sheet('Bookings by Hour and Day')
    wb.create_sheet('Bookings by Category')
    wb.create_sheet('No Shows')
    wb.create_sheet('Time Difference')
    wb.create_sheet('Popular Users')
    wb.create_sheet('Late Returns')
    wsBookingsByEquipment = wb['Number of Bookings by Equipment']
    wsTimeDifference = wb['Time Difference']
    wsPopularHours = wb['Most Popular Hours']
    wsPopularDays = wb['Most Popular Days']
    wsHourAndDay = wb['Bookings by Hour and Day']
    wsBookingCount = wb['Bookings by Category']
    wsNoShows = wb['No Shows']
    wsPopularUsers = wb['Popular Users']
    wsLateReturns = wb['Late Returns']

    #Number of Bookings by Equipment
    listsToExcel(wsBookingToEquipment, equipmentList, countList, 'Equipment Name', 'Number of Bookings', 35, 22)
    wsBookingToEquipment['D1'] = 'Total Number of Bookings'
    wsBookingToEquipment['D2'] = '=SUM(B2:B1000)'
    wsBookingToEquipment.column_dimensions['D'].width = 22

    #Most Popular Days
    for i in range(7):
        wsPopularDays['A' + str(i + 2)] = dayList[i]
        wsPopularDays['B' + str(i + 2)] = popularDayCount[i]
    setTotal(wsPopularDays, 'Total Number of Bookings', 8)
    wsPopularDays['A1'] = 'Day of the Week'
    wsPopularDays['B1'] = 'Number of Bookings' #incudes no-shows
    wsPopularDays.column_dimensions['A'].width = 15
    wsPopularDays.column_dimensions['B'].width = 15

    #Most Popular Hours
    for i in range(len(hourList)):
        hourList[i] = str(hourList[i]) + ':00'
    for i in range(len(hourList)):
        wsPopularHours['A' + str(i + 2)] = str(hourList[i]) + ':00'
        wsPopularHours['B' + str(i + 2)] = hourCountList[i]
    wsPopularHours.column_dimensions['A'].width = 8
    wsPopularHours.column_dimensions['B'].width = 12
    wsPopularHours['A1'] = 'Time'
    wsPopularHours['B1'] = 'Number of Bookings'
    setTotal(wsPopularHours, 'Total Number of Bookings', 25, 20)
    wsPopularHours.column_dimensions['B'].width = 22

    #Bookings by Hour and Day
    i = 0
    # for i in range(len(dayHourToCount)):
    #     dayHourToCount[i] = dayHourToCount[i] + ':00'
    for key in dayHourToCount:
        wsHourAndDay['A' + str(i + 2)] = dayListMon[key[0]] + ' ' + str(key[1]) + ':00'
        wsHourAndDay['B' + str(i + 2)] = dayHourToCount[key]
        i += 1
    wsHourAndDay['A1'] = 'Day and Time'
    wsHourAndDay['B1'] = 'Number of Bookings'
    setTotal(wsHourAndDay, 'Total Number of Bookings', 169)
    wsHourAndDay.column_dimensions['A'].width = 22
    wsHourAndDay.column_dimensions['B'].width = 22

    #Bookings by Category
    listsToExcel(wsBookingCount, categoryList, categoryCount, 'Category', 'Number of Bookings')
    setTotal(wsBookingCount, 'Total Number of Bookings', 100)
    wsBookingCount.column_dimensions['B'].width = 22

    #No Shows
    listsToExcel(wsNoShows, noShowUsername, noShowCount, 'Username', 'Number of No Shows')
    setTotal(wsNoShows, 'Total Number of No Shows', 1000)
    wsNoShows.column_dimensions['B'].width = 22

    #Time Difference
    timeDifferenceList.clear()
    timeDifferenceCount.clear()
    for i in range(len(timeDiffToCount)):
        if i > 2:
            wsTimeDifference['A' + str(i + 2)] = f'between {str(timeDiffToCount[i][0])} and {str(timeDiffToCount[i][0] + 1)} days'
            timeDifferenceList.append(f'between {str(timeDiffToCount[i][0])} and {str(timeDiffToCount[i][0] + 1)} days')
            timeDifferenceCount.append(timeDiffToCount[i][1])
        else:
            wsTimeDifference['A' + str(i + 2)] = timeDiffToCount[i][0]
            timeDifferenceList.append(timeDiffToCount[i][0])
            timeDifferenceCount.append(timeDiffToCount[i][1])
        wsTimeDifference['B' + str(i + 2)] = timeDiffToCount[i][1]


    setTotal(wsTimeDifference, 'Total Reservations', 100)
    wsTimeDifference['A2'] = 'Immediate Checkouts'
    wsTimeDifference['A3'] = '< 1 hour'
    wsTimeDifference['A4'] = 'between 1 hour and 1 day'
    wsTimeDifference['A1'] = 'Time Difference by Days'
    wsTimeDifference['B1'] = 'Number of Reservations'
    wsTimeDifference.column_dimensions['A'].width = 25

    #TODO THIS SHOULD BE CHECKED FOR CALCULATION ACCURACY AND SORTED

    #Popular Users
    namesKeys = list(nameToBookings.keys())
    equipmentAndBookings = list(nameToBookings.values())
    sortedPopularUsers = sortPair(namesKeys, equipmentAndBookings)
    equipmentAndBookings.sort(reverse=True)
    wsPopularUsers['A1'] = 'Username'
    wsPopularUsers['B1'] = 'Number of Equipment'
    wsPopularUsers['C1'] = 'Number of Unique Bookings'
    for i in range(len(sortedPopularUsers)):
        wsPopularUsers['A' + str(i + 2)] = sortedPopularUsers[i]
        wsPopularUsers['B' + str(i + 2)] = equipmentAndBookings[i][0]
        wsPopularUsers['C' + str(i + 2)] = equipmentAndBookings[i][1]
    setTotal(wsPopularUsers, 'Total Number of Equipment', 1000)
    wsPopularUsers.column_dimensions['A'].width = 20
    wsPopularUsers.column_dimensions['B'].width = 22
    wsPopularUsers.column_dimensions['C'].width = 25
    numberUniqueBookings = []
    numberEquipment = []
    for i in range(len(equipmentAndBookings)):
        numberUniqueBookings.append(equipmentAndBookings[i][0])
        numberEquipment.append(equipmentAndBookings[i][1])
    


    #Late Returns
    nameToLateReturn.sort(reverse=True)
    wsLateReturns['A1'] = 'Username'
    wsLateReturns['B1'] = 'Amount of Time Late'
    for i in range(len(nameToLateReturn)):
        wsLateReturns['A' + str(i + 2)] = nameToLateReturn[i][1]
        wsLateReturns['B' + str(i + 2)] = nameToLateReturn[i][2]
    wsLateReturns.column_dimensions['B'].width = 30
    print(nameToLateReturn)
    lateReturnMinutes = []
    lateReturnNames = []
    for i in range(len(nameToLateReturn)):
        lateReturnMinutes.append(int(nameToLateReturn[i][0] // 60))
        lateReturnNames.append(nameToLateReturn[i][1])

    #Sets zoom level for all worksheets
    for sheet in wb.worksheets:
        sheet.sheet_view.zoomScale = 300

    #Save the Worbook
    wb.save('bookingAnalysis.xlsx')

    dayHourListChart = []
    for dayHour in dayHourList:
        if dayHour[0] == 0:
            dayHourListChart.append(f'Monday {dayHour[1]}:00')
        elif dayHour[0] == 1:
            dayHourListChart.append(f'Tuesday {dayHour[1]}:00')
        elif dayHour[0] == 2:
            dayHourListChart.append(f'Wednesday {dayHour[1]}:00')
        elif dayHour[0] == 3:
            dayHourListChart.append(f'Thursday {dayHour[1]}:00')
        elif dayHour[0] == 4:
            dayHourListChart.append(f'Friday {dayHour[1]}:00')
        elif dayHour[0] == 5:
            dayHourListChart.append(f'Saturday {dayHour[1]}:00')
        elif dayHour[0] == 6:
            dayHourListChart.append(f'Sunday {dayHour[1]}:00')

    print(timeDifferenceList)
    timeDifferenceList[0] = 'Immediate Checkout'
    timeDifferenceList[1] = '< 1 hour'
    timeDifferenceList[2] = 'between 1 hour and 1 day'
   
    return (equipmentList, countList, dayList, popularDayCount, hourList, hourCountList, dayHourListChart, 
        list(dayHourToCount.values()), categoryList, categoryCount, noShowUsername, noShowCount, timeDifferenceList, timeDifferenceCount,
        sortedPopularUsers, numberEquipment, numberUniqueBookings, lateReturnNames, lateReturnMinutes)

#3454 is the total number of items booked (Number of Bookings by Equipment Total, Bookings by Category)
#1985 is the total number of unique bookings (Bookings by Hour and , most popular hours, Most Popular Days, ) 
#1390 - Time difference total does not add up to either of these 